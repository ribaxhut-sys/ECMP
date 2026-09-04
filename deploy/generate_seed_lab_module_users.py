#!/usr/bin/env python3
"""Generate lab directory: 200 calon identitas — SEMUA belum terdaftar.

  - username = ID login pendek (3101, 3102, …)
  - email / external_user_id = ID 16 digit identitas pusat
  - display_name = nama orang
  - TIDAK ada peran — admin memilih peran saat mendaftarkan
  - home_branch = lokasi referensi (boleh diubah admin)
"""

from __future__ import annotations

import csv
import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUT_DIR = Path(__file__).resolve().parent
NS = uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890")

# Official UPPPD — sync with alembic 0047 (42 rows in live DB)
UPPPD: list[tuple[str, str]] = [
    ("Jakarta Pusat", "UPPPD Tanah Abang"),
    ("Jakarta Pusat", "UPPPD Gambir"),
    ("Jakarta Pusat", "UPPPD Sawah Besar"),
    ("Jakarta Pusat", "UPPPD Kemayoran"),
    ("Jakarta Pusat", "UPPPD Senen"),
    ("Jakarta Pusat", "UPPPD Cempaka Putih"),
    ("Jakarta Pusat", "UPPPD Menteng"),
    ("Jakarta Pusat", "UPPPD Johar Baru"),
    ("Jakarta Utara", "UPPPD Penjaringan"),
    ("Jakarta Utara", "UPPPD Pademangan"),
    ("Jakarta Utara", "UPPPD Tanjung Priok"),
    ("Jakarta Utara", "UPPPD Koja"),
    ("Jakarta Utara", "UPPPD Kelapa Gading"),
    ("Jakarta Utara", "UPPPD Cilincing"),
    ("Jakarta Barat", "UPPPD Grogol Petamburan"),
    ("Jakarta Barat", "UPPPD Taman Sari"),
    ("Jakarta Barat", "UPPPD Tambora"),
    ("Jakarta Barat", "UPPPD Kebon Jeruk"),
    ("Jakarta Barat", "UPPPD Palmerah"),
    ("Jakarta Barat", "UPPPD Kembangan"),
    ("Jakarta Barat", "UPPPD Cengkareng"),
    ("Jakarta Barat", "UPPPD Kalideres"),
    ("Jakarta Selatan", "UPPPD Kebayoran Baru"),
    ("Jakarta Selatan", "UPPPD Kebayoran Lama"),
    ("Jakarta Selatan", "UPPPD Pesanggrahan"),
    ("Jakarta Selatan", "UPPPD Cilandak"),
    ("Jakarta Selatan", "UPPPD Pasar Minggu"),
    ("Jakarta Selatan", "UPPPD Jagakarsa"),
    ("Jakarta Selatan", "UPPPD Mampang Prapatan"),
    ("Jakarta Selatan", "UPPPD Pancoran"),
    ("Jakarta Selatan", "UPPPD Tebet"),
    ("Jakarta Selatan", "UPPPD Setiabudi"),
    ("Jakarta Timur", "UPPPD Matraman"),
    ("Jakarta Timur", "UPPPD Pulogadung"),
    ("Jakarta Timur", "UPPPD Jatinegara"),
    ("Jakarta Timur", "UPPPD Duren Sawit"),
    ("Jakarta Timur", "UPPPD Kramat Jati"),
    ("Jakarta Timur", "UPPPD Makasar"),
    ("Jakarta Timur", "UPPPD Pasar Rebo"),
    ("Jakarta Timur", "UPPPD Ciracas"),
    ("Jakarta Timur", "UPPPD Cipayung"),
    ("Jakarta Timur", "UPPPD Cakung"),
]

# Four people per UPPPD as a directory pool only — role is NOT assigned here.
PEOPLE_PER_BRANCH = 4

FIRST_NAMES = [
    "Ahmad", "Budi", "Citra", "Dewi", "Eko", "Fitri", "Galih", "Hana",
    "Indra", "Joko", "Kartika", "Lestari", "Maya", "Nanda", "Oki", "Putri",
    "Rudi", "Sari", "Tono", "Umi", "Vina", "Wawan", "Yuni", "Zaki",
    "Agus", "Bayu", "Dian", "Fajar", "Gilang", "Hendra", "Intan", "Jihan",
    "Kirana", "Lukman", "Mega", "Nurul", "Prasetyo", "Rina", "Surya", "Tari",
]

LAST_NAMES = [
    "Santoso", "Wijaya", "Kusuma", "Saputra", "Lestari", "Rahayu", "Nugroho",
    "Hidayat", "Putra", "Putri", "Sari", "Wibowo", "Gunawan", "Halim",
    "Firmansyah", "Mahendra", "Permata", "Anggraini", "Syahputra", "Nasution",
    "Simanjuntak", "Siregar", "Harahap", "Manurung", "Panggabean", "Situmorang",
    "Hutagalung", "Lubis", "Dalimunthe", "Pratama", "Kurniawan", "Setiawan",
    "Handayani", "Safitri", "Maulana", "Ramadhan", "Oktaviani", "Prasetya",
    "Cahyadi", "Susanti",
]


def branch_code(name: str) -> str:
    return name.upper().replace(" ", "-")


def uid(*parts: str) -> uuid.UUID:
    return uuid.uuid5(NS, "|".join(parts))


def person_id(seq: int) -> str:
    """16-digit central identity (email local-part / external_user_id)."""
    return f"{3100000000000000 + seq:016d}"


def login_username(seq: int) -> str:
    """Short lab login: 3100000000000001 → 3101."""
    return f"31{seq:02d}"


def person_name(seq: int) -> str:
    """Deterministic Indonesian-style name; uniqueness enforced in build_rows."""
    i = seq - 1
    first = FIRST_NAMES[(i * 3) % len(FIRST_NAMES)]
    last = LAST_NAMES[(i * 7) % len(LAST_NAMES)]
    return f"{first} {last}"


def unique_person_name(seq: int, used: set[str]) -> str:
    base = person_name(seq)
    if base not in used:
        return base
    i = seq - 1
    for offset in range(1, 200):
        first = FIRST_NAMES[(i * 3 + offset) % len(FIRST_NAMES)]
        last = LAST_NAMES[(i * 7 + offset * 5) % len(LAST_NAMES)]
        candidate = f"{first} {last}"
        if candidate not in used:
            return candidate
    raise SystemExit(f"cannot allocate unique name for seq={seq}")


def build_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seq = 0
    used_names: set[str] = set()

    # Identity directory only — no role. Admin assigns role at registration.
    for city, name in UPPPD:
        code = branch_code(name)
        for _slot in range(PEOPLE_PER_BRANCH):
            seq += 1
            user_id = person_id(seq)
            full = unique_person_name(seq, used_names)
            used_names.add(full)
            rows.append(
                {
                    "username": login_username(seq),
                    "display_name": full,
                    "email": f"{user_id}@lab.ecmp.local",
                    "external_user_id": user_id,
                    "home_branch_code": code,
                    "home_branch_name": name,
                    "region": city,
                    "registration_status": "PENDING",
                    "cohort": "DIRECTORY_POOL",
                    "notes": (
                        "Calon identitas pusat — BELUM anggota modul. "
                        "Peran ditentukan admin saat mendaftarkan."
                    ),
                    "candidate_id": str(uid("candidate", user_id)),
                }
            )

    for _ in range(32):
        seq += 1
        user_id = person_id(seq)
        city, name = UPPPD[(seq - 1) % len(UPPPD)]
        full = unique_person_name(seq, used_names)
        used_names.add(full)
        rows.append(
            {
                "username": login_username(seq),
                "display_name": full,
                "email": f"{user_id}@lab.ecmp.local",
                "external_user_id": user_id,
                "home_branch_code": branch_code(name),
                "home_branch_name": name,
                "region": city,
                "registration_status": "PENDING",
                "cohort": "EXTRA_CANDIDATE",
                "notes": (
                    "Calon ekstra — boleh tidak didaftarkan; "
                    "uji user pusat ada tapi belum anggota modul."
                ),
                "candidate_id": str(uid("candidate", user_id)),
            }
        )

    assert len(UPPPD) == 42, len(UPPPD)
    assert len(rows) == 200, len(rows)
    assert all(r["username"] == login_username(i) for i, r in enumerate(rows, start=1))
    assert len({r["username"] for r in rows}) == 200
    assert len({r["display_name"] for r in rows}) == 200
    return rows


COLUMNS = [
    "username",
    "display_name",
    "email",
    "external_user_id",
    "home_branch_code",
    "home_branch_name",
    "region",
    "registration_status",
    "cohort",
    "notes",
    "candidate_id",
]


def write_csv(rows: list[dict[str, str]], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "candidates"
    ws.append(COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
    # Keep username as text (leading zeros / Excel number mangling)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "@"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    readme = wb.create_sheet("README", 0)
    for line in (
        "ECMP Lab — Daftar CALON identitas (Mode A)",
        "",
        "username = ID login pendek (3101, 3102, …)",
        "email / external_user_id = ID 16 digit identitas pusat",
        "display_name = nama orang",
        "",
        "PERAN TIDAK diisi di directory — admin memilih peran saat mendaftarkan.",
        "home_branch_* = lokasi kerja referensi (boleh diubah admin).",
        "",
        "200 PENDING: 168 pool (42 UPPPD × 4 orang) + 32 ekstra.",
        "Login: 3101 … 31200 (identity 16 digit tetap di email/external_user_id)",
    ):
        readme.append([line])
    wb.save(path)


def write_sql_cleanup(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "-- Cleanup only — calon TIDAK di-insert; admin mendaftarkan.",
                "BEGIN;",
                "DELETE FROM user_roles WHERE user_id IN (",
                "  SELECT id FROM users WHERE email LIKE '%@lab.ecmp.local'",
                ");",
                "DELETE FROM refresh_tokens WHERE user_id IN (",
                "  SELECT id FROM users WHERE email LIKE '%@lab.ecmp.local'",
                ");",
                "DELETE FROM password_reset_tokens WHERE user_id IN (",
                "  SELECT id FROM users WHERE email LIKE '%@lab.ecmp.local'",
                ");",
                "DELETE FROM users WHERE email LIKE '%@lab.ecmp.local';",
                "COMMIT;",
                "",
            ]
        ),
        encoding="utf-8",
    )


def write_frontend_json(rows: list[dict[str, str]], path: Path) -> None:
    """Compact candidate list for FE register-user search (Mode A lab)."""
    import json

    payload = [
        {
            "username": r["username"],
            "displayName": r["display_name"],
            "email": r["email"],
            "homeBranchCode": r["home_branch_code"],
            "homeBranchName": r["home_branch_name"],
            "region": r["region"],
            "cohort": r["cohort"],
        }
        for r in rows
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    rows = build_rows()
    stem = OUT_DIR / "seed-lab-module-users-200"
    write_csv(rows, Path(f"{stem}.csv"))
    write_xlsx(rows, Path(f"{stem}.xlsx"))
    write_sql_cleanup(Path(f"{stem}.sql"))
    fe_json = (
        OUT_DIR.parent
        / "frontend"
        / "src"
        / "features"
        / "users"
        / "data"
        / "moduleUserCandidates.json"
    )
    write_frontend_json(rows, fe_json)
    pending = sum(1 for r in rows if r["registration_status"] == "PENDING")
    pool = sum(1 for r in rows if r["cohort"] == "DIRECTORY_POOL")
    extra = sum(1 for r in rows if r["cohort"] == "EXTRA_CANDIDATE")
    print(
        f"Wrote {stem}.{{csv,xlsx,sql}} + {fe_json} — "
        f"PENDING={pending} POOL={pool} EXTRA={extra} "
        f"sample: {rows[0]['username']} / {rows[0]['display_name']}"
    )


if __name__ == "__main__":
    main()
